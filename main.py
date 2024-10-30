from  openpyxl import Workbook, load_workbook
import gspread
from google.oauth2.service_account import Credentials


from secret_variables import sheet_id

#=======================================================================================
# Declaration of arraies for the openPyXL.
# Needs to be here for the Gspread for loop to work
#=======================================================================================

# test arry
TestArry = [['02-05-2023','Jumbo','24.86'],['02-05-2023','Jumbo','8'],['05-05-2023','Jumbo','61.46']]

#==============
# Arrays for production
Salery = []
HealthInsur = [] 
House = [] # rent mainly other bank transactions
HouseGarCloth = []
EatingOut = []
groceries = []
Uni = [] # need to add to category detection helper
Cats = [] # need to add to category detection helper
Garden = [] # need to add to category detection helper
Gym = []
PreserntsBD = []
EverythingElse = [] # everything else will be added to this arry, to viewing and manual work

#=======================================================================================

# cell locaiton for each category to start updating from

SaleryIndex = 'B7'
HealthInsurIndex = 'B12'
HouseIndex = 'B15' # rent mainly other bank transactions
HouseGarClothIndex = 'B33'
EatingOutIndex = 'J5'
groceriesIndex = 'F5'
UniIndex = 'N5' # need to add to category detection helper
CatsIndex = 'N15' # need to add to category detection helper
GardenIndex = 'N24'
GymIndex = 'N34'
PreserntsBDIndex = 'B36'
EverythingElseIndex = 'R5'



#=======================================================================================
# Gspread and GCP credantials ==========================================================
#=======================================================================================

scopes = ["https://www.googleapis.com/auth/spreadsheets"]
creds = Credentials.from_service_account_file("./secret_variables/credentials.json", scopes=scopes)
gc = gspread.authorize(creds)

# sheet_id = "1qDbyPXXKwC9VRDh40Ej2d-87XX6-Rabscz3nQCBBpGo"
sh = gc.open_by_key(sheet_id)


# Worksheets:
worksheetMain = sh.worksheet("main-2023")
worksheetSepDec = sh.worksheet("Detailed  Sep-Dec")
worksheetTest = sh.worksheet("Test")
#values_list = worksheetMain.row_values(2)
#print(values_list)

#worksheetTest.update_acell("A1", "Lorem ipsum")

#============================================================================================
# Finding a month's name in the google sheet, and itereting it's cell's column with 2 variables, Row and Col.
# For example: "September" is in cell B2, so column B needs to be itterated with the next 2 cells to it's right, 
# So B, C, D (date, description and amount).
# Get the Month's row  and col Col = monthA.col, same for Row, in this case, row = 2, col = b (2)
# Why we do this? if you take a look at the referenced pictures, you can see that in one sheet there are 4 months
# next to each other, under each we have the expences listed.
# update Row + 1 (to skip the title) have it in a loop as i and iterate it
#============================================================================================

MonthA = worksheetTest.find("MonthA") 


# Figure this out: ==============
# Row = MonthA.row
# Col = MonthA.col
# ===============================


# print at Row+1
#worksheetTest.update_cell(Row+1, Col, "Something else!")


# # Find in a column
# CatLocationHouse = worksheetTest.find("House",None,Col,True)
# CatHouseRow = CatLocationHouse.row
# CatHouseCol = CatLocationHouse.col
# CatUpdateCell = (CatLocationHouse.row+1,CatLocationHouse.col)



# CellInFourRows = worksheetTest.cell(CatHouseRow + 4, CatHouseCol)


# while (CellInFourRows.value is not None):
    # worksheetTest.update_cell(CatHouseRow + 1, CatHouseCol,"This is a test2")
    # CatHouseRow +=1

# For batch_update ======================================================

# CurrRow = CatHouseRow # to have the row number as an iterable index in the loop
# CellsToUpdate = worksheetTest.cell(CatLocationHouse.row,CatLocationHouse.col)
# CellList = [[CurrRow,CatHouseCol],[CurrRow,CatHouseCol+1],[CurrRow,CatHouseCol+2]]
# # print(CellList)
NewCellListSTR = worksheetTest.range('B12:D15')

#============================================================================================
# batch_update WORKS!! without a loop, the cell in the 'range' is the first cell to update, and it wili update 
# depending on the 'majorDimension' untill the array is done.
# need to look int adding batch update into a if - to add empty rows before overwriting the whole gsheet
#============================================================================================

# worksheetTest.update_cells(NewCellListSTR,TestArry)

# cell_list = worksheetTest.range('A1:C1')

# for i in TestArry:
#     print(i.value)
#     worksheetTest.update_cell(CatLocationHouse.row, CatLocationHouse.col, TestArry[i])
#     print("I is: ", i)
#     print(i.value)

#============================================================================================
# Since batch update wont work with cell object, the locations of the tables for each category will be 
# Fixed and be declared at the top, like this:
# GroceriesIndex = 'F5'
#============================================================================================


# GroceriesIndexTest = 'F5'

# worksheetTest.batch_update([{
#     'range': GroceriesIndexTest,
#     'values': TestArry,
#     'majorDimension': 'ROWS'
# }])

# ===================================================================

# Testing update method since loop in not working - "Invalid value at 'data.values[1]' (type.googleapis.com/google.protobuf.ListValue), "Jumbo""

# worksheetTest.update()

#MonthCell = worksheetSepDec.find("September") # find September month, to assign arrays of month here
#print(MonthCell.row, MonthCell.col)



#CatCellEating = worksheetSepDec.find("eating out")
#print("Row is: ", CatCellEating.row," Col is: ", CatCellEating.col) # find a specific category for the arrays


#=======================================================================================
# OpenPyXL declaration and info ========================================================
#=======================================================================================

wb = load_workbook('./Test.xlsx')
ws = wb.active # the work sheet (Tab) we are working on currently
# #print(ws)


i = 2 # manualk set for i for testing and the start of the loop
cellindxD = "D" + str(i) # insurting the value into a cell formula: D2, D3, D10 etc.

# cell is the current cell of D and i which will be used in a loop to go over all D cells.
cellD = str(ws[cellindxD].value)


# Creating new workbook, the while loop will add info to it
newWB = Workbook()
newWS = newWB.active
newWS["A1"].value = "Date"
newWS["B1"].value = "Description"
newWS["C1"].value = "Value"


#=======================================================================================
# Hellper func - Categories
#=======================================================================================


def category(cellindxH):
    CellCateg = str(ws[cellindxH].value)

    #print("Test cell is not empty")
    #print(cellindxH)
    #print(CellCateg)
    

    if (CellCateg.find("CZ Groep") != -1):
        newWS[newCellIndexB].value = "CZ Insurance"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        HealthInsur.append(Thisisarow)

    elif (CellCateg.find("  IBAN: NL") != -1):
        newWS[newCellIndexB].value = "Bank Transaction"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        House.append(Thisisarow)

    elif ((CellCateg.find("Salaris") != -1) or (CellCateg.find("Salary ") != -1)):
        newWS[newCellIndexB].value = "Salery"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        Salery.append(Thisisarow)

    elif (CellCateg.find("GREEN FIT") != -1):
        newWS[newCellIndexB].value = "Green Fit gym"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        Gym.append(Thisisarow)

    # groceries =======================================================================================
    elif (CellCateg.find("Jumbo") != -1):
        newWS[newCellIndexB].value = "Jumbo"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        groceries.append(Thisisarow)
        print(groceries)

    elif ((CellCateg.find("Albert Heijn ") != -1) or (CellCateg.find("ALBERT HEIJN ") != -1)):
        newWS[newCellIndexB].value = "Albert Heijn"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        groceries.append(Thisisarow)
    
    # Eating out, Restorants, Starbucks etc: ===========================================================
    elif (CellCateg.find("STARBUCKS,") != -1):
        newWS[newCellIndexB].value = "Starbucks"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        EatingOut.append(Thisisarow)

    elif (CellCateg.find("   Set Genki Tei") != -1):
        newWS[newCellIndexB].value = "SET"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        EatingOut.append(Thisisarow)

    elif (CellCateg.find("   The Avocado Show ") != -1):
        newWS[newCellIndexB].value = "Avocado show"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        EatingOut.append(Thisisarow)

    elif (CellCateg.find("   MAAS - Universiteit ") != -1):
        newWS[newCellIndexB].value = "Uni coffee"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        EatingOut.append(Thisisarow)

    # House / garden / cloths/ self care: ===========================================================
    elif ((CellCateg.find("  MM ") != -1) or (CellCateg.find("MediaMarkt ") != -1)):
        newWS[newCellIndexB].value = "Media Market"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        HouseGarCloth.append(Thisisarow)

    elif (CellCateg.find("     HEMA ") != -1):
        newWS[newCellIndexB].value = "Hema"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        HouseGarCloth.append(Thisisarow)
    
    elif (CellCateg.find("   Action ") != -1):
        newWS[newCellIndexB].value = "Action"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        HouseGarCloth.append(Thisisarow)

    elif (CellCateg.find(" ETOS ") != -1):
        newWS[newCellIndexB].value = "Etos"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        HouseGarCloth.append(Thisisarow)

    elif (CellCateg.find("  HOFSTEDE,") != -1):
        newWS[newCellIndexB].value = "HOFSTEDE"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        HouseGarCloth.append(Thisisarow)

    elif (CellCateg.find("   H&M ") != -1):
        newWS[newCellIndexB].value = "H&M"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        HouseGarCloth.append(Thisisarow)
    
    elif (CellCateg.find("Dam Apotheek,") != -1): # pharmacy is part of "House / self care"
        newWS[newCellIndexB].value = "Pharmacy"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        HouseGarCloth.append(Thisisarow)

    # Presents Bday and hbbies =============================================================================
    elif (CellCateg.find("Pipoos Den Haag") != -1): 
        newWS[newCellIndexB].value = "Pipoos"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        PreserntsBD.append(Thisisarow)

    elif (CellCateg.find("Airbnb") != -1): 
        newWS[newCellIndexB].value = "Airbnb"
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        PreserntsBD.append(Thisisarow)

    else:
        print("I didnt find this word")
        newWS[newCellIndexB].value = CellCateg
        Thisisarow = [newWS[newCellIndexA].value, newWS[newCellIndexB].value, newWS[newCellIndexC].value]
        EverythingElse.append(Thisisarow)
        

#=======================================================================================
# A function that batch updates the Gsheet with the Caterotty arrys from the excel.
#=======================================================================================

def updatingGsheeWithCat():
    worksheetTest.batch_update([{
        'range': groceriesIndex,
        'values': groceries,
        'majorDimension': 'ROWS'
        }])

    worksheetTest.batch_update([{
        'range': EatingOutIndex,
        'values': EatingOut,
        'majorDimension': 'ROWS'
        }])

    worksheetTest.batch_update([{
        'range': EverythingElseIndex,
        'values': EverythingElse,
        'majorDimension': 'ROWS'
        }])
    
    worksheetTest.batch_update([{
        'range': SaleryIndex,
        'values': Salery,
        'majorDimension': 'ROWS'
        }])
    
    worksheetTest.batch_update([{
        'range': HealthInsurIndex,
        'values': HealthInsur,
        'majorDimension': 'ROWS'
        }])




#=======================================================================================
# The while loop that should work in the end.
# it takes the value in the cell and sepoerationg it into year month and day, 
# that rearanges them into a new value that will be put into a cell.
#=======================================================================================


#start of while loop

while (ws[cellindxD].value is not None):
    cellindxD = "D" + str(i) # Col D is dates in original sheet
    cellindxG = "G" + str(i) # Col G is value in original sheet
    cellindxH = "H" + str(i) # Col H is Description, and will be set to Col B in new sheet

    newCellIndexA = "A" + str(i) # creating a new index for col A in new sheet, For dates from D
    newCellIndexB = "B" + str(i) # creating a new index for col B in new sheet For Value from H
    newCellIndexC = "C" + str(i) # creating a new index for col C in new sheet For Value from G

    # getting date from col D, changing it and insurting into col A =============================
    cellD = str(ws[cellindxD].value)
    #print("This is the cell value: " + cellD)
    year = cellD[0:4]
    month = cellD[4:6]
    day = cellD[6:]
    newDate = year + "-" + month + "-" + day
    newWS[newCellIndexA].value = newDate
    #print(newWS[newCellIndexA].value) # printing cells from new sheet with new date
    print("")

    # getting value from col G and inserting to col C ===========================================
    cellG = ws[cellindxG].value
    try:
        cellG = -1*(int(cellG))
    except Exception:
        pass

        #print(cellG)
    newWS[newCellIndexC].value = cellG

    # setting category, if no found - paste description =======================================
    if ws[cellindxH].value is None:
        print("Test cell is empty!")
        # break ## Dont think this is how it works, just wanted to tell myself what to do here
    else:
        category(cellindxH)
    
    # Saving excel sheet since we are done updating it at this point ==========================
    try:
        # Save the workbook
        newWB.save("C:/Users/shori/OneDrive/Desktop/Code_projects/Py-Sheets/NewDocResult.xlsx")
        print("Changes saved successfully!")
    except Exception as e:
        print("Error saving changes:", e)

    i+=1
 # End of while loop! =================================================================================

updatingGsheeWithCat()
print("This is after calling the update function")

# Adding the category arrays to the Gsheet by using batch update, outside of while loop that created new Excel ============================


print("Another random print")


#=======================================================================================
# Trying to find categories
#=======================================================================================


# i = 1
# cellindxH = "H" + str(i)

# JumboCount = 0
# AHCount = 0
# CZCount = 0
# StarbucksCount = 0

 
# while (ws[cellindxH].value is not None):
#     cellindxH = "H" + str(i)
#     testCellCateg = str(ws[cellindxH].value)
#     print(i)
#     if ws[cellindxH].value is None:
#         print("Test cell is empty!")
#         break
#     else:
#         print("Test cell is not empty")
#         print(cellindxH)
#         print(testCellCateg)
#         if (testCellCateg.find("CZ Groep") != -1):
#             print("I found what you were looking for!")
#             CZCount +=1
#             i += 1
#         elif (testCellCateg.find("Jumbo") != -1):
#             print("I found Jumbo!")
#             JumboCount +=1
#             i += 1
#         elif ((testCellCateg.find("Albert Heijn ") != -1) or (testCellCateg.find("ALBERT HEIJN ") != -1)):
#             print("I found Albert Heijn !")
#             AHCount +=1
#             i += 1
#         elif (testCellCateg.find("STARBUCKS,") != -1):
#             print("I found Starbucks!")
#             StarbucksCount +=1
#             i += 1
#         else:
#             print("I didnt find this word")
#             i += 1

# print("JumboCount is:",JumboCount, " AHCount is:", AHCount, " CZCount is:",CZCount, " StarbucksCount is:", StarbucksCount)





#=======================================================================================
# Test with try and except to save changes to cell
#=======================================================================================

# print(cell)
# cell = "This is a test"
# wb.template = False
# try:
#     # Save the workbook
#     newWB = workbook()
#     newWS = newWB.active    
#     newWB.save("C:/Users/shori/OneDrive/Desktop/Code_projects/Py-Sheets/Test3.xlsx")
#     print("Changes saved successfully!")
# except Exception as e:
#     print("Error saving changes:", e)

# # wb.save('./Test.xlsx')
# print(cell)



#=======================================================================================
# Test cell for one change in manually set cell
#=======================================================================================

# testCell = str(ws["D215"].value)

# if ws["D216"].value is None:
#     print("Test cell is empty!")
#     ws["D216"].value = "I am a test cell!"
#     print(ws["D216"].value)
#     wb.save('Test.xlsx')
# else:
#     print("Test cell is not empty")










#=======================================================================================
# Initian plan and goals
#=======================================================================================
# For loop, For cell 1:
# year = string(0, 4) not including 4
# month = string(4, 6) not including 6
# day = string(6, , ) 6 onwards.
# return (year, "-",month, "-", day)
# re-write the current cell into new value
# continue to next row.

# Save all changes to wb, or save after every row?? 

#=======================================================================================
# how to excecute the date rearange
#=======================================================================================

# print(type(cell))
# year = cell[0:4]
# print(year)

# month = cell[4:6]
# print(month)

# day = cell[6:]
# print(day)

# newDate = year + "-" + month + "-" + day
# print(newDate)


#=======================================================================================
# ChatGPT code suggestion
#=======================================================================================
# from openpyxl import load_workbook

# try:
#     # Load existing workbook
#     wb = load_workbook("Test2.xlsx")
    
#     # Select worksheet
#     ws = wb.active
    
#     # Modify cell values
#     ws['A1'] = 'New Value'
#     ws['B2'] = 123
    
#     # Save the workbook
#     wb.save("Test2.xlsx")
#     print("Changes saved successfully!")
# except Exception as e:
#     print("Error saving changes:", e)


# try:
#     # Create a new workbook
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Grades"
    
#     # Modify cell values
#     ws['A1'] = 'Hello'
#     ws['B1'] = 'World'
    
#     # Save the workbook
#     wb.save("example.xlsx")
#     print("File saved successfully!")
# except Exception as e:
#     print("Error saving file:", e)